from flask import Flask, render_template, request, redirect, send_file, session, url_for
from datetime import datetime, timedelta
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

app = Flask(__name__)
app.secret_key = "super_secret_safety_key" # In production, use an environment variable

# ------------------------
# BASE DE DATOS
# ------------------------

def init_db():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS contador (
        id INTEGER PRIMARY KEY,
        fecha_inicio TEXT,
        record_dias INTEGER
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS historial (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_inicio TEXT,
        fecha_fin TEXT,
        dias INTEGER,
        horas INTEGER,
        minutos INTEGER,
        segundos INTEGER
    )
    """)

    cursor.execute("SELECT * FROM contador WHERE id = 1")

    if cursor.fetchone() is None:
        cursor.execute(
            "INSERT INTO contador (id, fecha_inicio, record_dias) VALUES (1, ?, 0)",
            (datetime.now().isoformat(),)
        )

    conn.commit()
    conn.close()


def obtener_record():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT record_dias FROM contador WHERE id = 1")
    record = cursor.fetchone()[0]
    conn.close()
    return record


def actualizar_record(dias):
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("UPDATE contador SET record_dias = ? WHERE id = 1", (dias,))
    conn.commit()
    conn.close()


def obtener_fecha():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT fecha_inicio FROM contador WHERE id = 1")
    fecha = cursor.fetchone()[0]
    conn.close()
    return datetime.fromisoformat(fecha)


def reiniciar():
    fecha_inicio = obtener_fecha()
    ahora = datetime.now()
    diferencia = ahora - fecha_inicio

    dias = diferencia.days
    segundos_totales = diferencia.seconds
    horas = segundos_totales // 3600
    minutos = (segundos_totales % 3600) // 60
    segundos = segundos_totales % 60

    record_actual = obtener_record()
    if dias > record_actual:
        actualizar_record(dias)

    # GUARDAR EN HISTORIAL
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO historial (fecha_inicio, fecha_fin, dias, horas, minutos, segundos)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        fecha_inicio.isoformat(),
        ahora.isoformat(),
        dias,
        horas,
        minutos,
        segundos
    ))

    # REINICIAR CONTADOR
    cursor.execute("UPDATE contador SET fecha_inicio = ? WHERE id = 1",
                   (ahora.isoformat(),))

    conn.commit()
    conn.close()


def borrar_historial():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("DELETE FROM historial")
    cursor.execute("UPDATE contador SET record_dias = 0 WHERE id = 1")
    # Opcional: reiniciar también la fecha de inicio al borrar todo
    cursor.execute("UPDATE contador SET fecha_inicio = ? WHERE id = 1", (datetime.now().isoformat(),))
    conn.commit()
    conn.close()


# ------------------------
# EXPORTAR EXCEL
# ------------------------

@app.route("/exportar")
def exportar_excel():
    if not session.get("admin"):
        return redirect(url_for("admin"))
        
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT fecha_inicio, fecha_fin, dias, horas, minutos, segundos FROM historial ORDER BY id DESC")
    datos = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Accidentes"

    # Estilos para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    alignment = Alignment(horizontal="center")

    headers = ["Fecha Inicio", "Fecha Fin", "Días", "Horas", "Minutos", "Segundos"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    for fila in datos:
        # Formatear fechas para que sean más legibles en Excel
        f_inicio = datetime.fromisoformat(fila[0]).strftime("%Y-%m-%d %H:%M:%S")
        f_fin = datetime.fromisoformat(fila[1]).strftime("%Y-%m-%d %H:%M:%S")
        ws.append([f_inicio, f_fin, fila[2], fila[3], fila[4], fila[5]])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    archivo = "historial_accidentes.xlsx"
    wb.save(archivo)

    return send_file(archivo, as_attachment=True)

@app.route("/reiniciar_accidente", methods=["POST"])
def ruta_reiniciar():
    if not session.get("admin"):
        return redirect(url_for("admin"))

    reiniciar()
    return redirect(url_for("admin"))


@app.route("/limpiar_historial", methods=["POST"])
def ruta_limpiar():
    if not session.get("admin"):
        return redirect(url_for("admin"))

    borrar_historial()
    return redirect(url_for("admin"))


# ------------------------
# RUTA PRINCIPAL
# ------------------------

@app.route("/")
def home():
    fecha_inicio = obtener_fecha()
    return render_template("index.html", 
                         fecha_inicio=fecha_inicio.isoformat(),
                         record=obtener_record())


# ------------------------
# ADMIN
# ------------------------

@app.route("/login", methods=["POST"])
def login():
    if request.form.get("password") == "admin123":
        session["admin"] = True
    return redirect(url_for("admin"))


@app.route("/logout")
def logout():
    session.pop("admin", None)
    return redirect(url_for("home"))


@app.route("/admin")
def admin():
    if not session.get("admin"):
        return render_template("login.html")
    
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM historial ORDER BY id DESC LIMIT 10")
    historial = cursor.fetchall()
    conn.close()
    
    return render_template("admin.html", historial=historial)


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
